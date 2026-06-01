"""
TrustPulse — pipeline/ingest/beds.py
Ingests NHS bed availability and occupancy data from two sources:

1. Monthly bed sitrep files (data/raw/beds/sitrep/)
   - April 2022 to March 2024: Excel format (.xlsx)
     Sheet: 'all acute trusts'
     National/region summary: rows 16-24
     Second header row: row 26
     Trust-level data: row 27 onwards
     Columns: Region, Trust Name, Code, G&A beds available, G&A beds occupied,
               G&A occupancy rate, Adult/Paediatric/Critical care breakdowns
   - April 2024 onwards: CSV format (.csv)
     Long/tidy format: one row per org per metric
     Level column: Provider = trust-level, Region, ICB, National also present
     Columns: Period, Level, Region, ICB, Org Code, Org Name, Metric, Type, Value

2. Quarterly KH03 bed occupancy files (data/raw/beds/kh03/)
   - Excel format (.xlsx)
   - Sheet: 'NHS Trust by Sector'
   - Metadata rows 1-14, headers at rows 14-15, data from row 16
   - Three metric groups: Available, Occupied, % Occupied
   - Sectors: Total, General & Acute, Learning Disability/Disabilities,
              Maternity, Mental Illness

Outputs:
    data/processed/beds_sitrep_clean.csv  — monthly trust-level bed metrics
    data/processed/beds_kh03_clean.csv    — quarterly trust-level bed occupancy by sector

Usage:
    python pipeline\ingest\beds.py
"""

import os
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# ── Paths ─────────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SITREP_DIR = os.path.join(BASE_DIR, "data", "raw", "beds", "sitrep")
KH03_DIR = os.path.join(BASE_DIR, "data", "raw", "beds", "kh03")
PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
SITREP_OUTPUT = os.path.join(PROCESSED_DIR, "beds_sitrep_clean.csv")
KH03_OUTPUT = os.path.join(PROCESSED_DIR, "beds_kh03_clean.csv")

# Sitrep metrics to retain from both formats
# These are the core G&A and critical care metrics relevant to TrustPulse
SITREP_METRICS_TO_KEEP = [
    "G&A beds available",
    "G&A beds occupied",
    "G&A occupancy rate",
    "Adult G&A beds available",
    "Adult G&A beds occupied",
    "Adult G&A occupancy rate",
    "Paediatric G&A beds available",
    "Paediatric G&A beds occupied",
    "Paediatric G&A occupancy rate",
    "Adult critical care beds available",
    "Adult critical care beds occupied",
    "Adult critical care occupancy rate",
    # Length of stay metrics (CSV format only, April 2024 onwards)
    "Number of G&A beds occupied by patients with a length of stay of - 7 or more days",
    "Number of G&A beds occupied by patients with a length of stay of - 14 or more days",
    "Number of G&A beds occupied by patients with a length of stay of - 21 or more days",
    "% occupied G&A beds occupied by patients with a length of stay of - 7 or more days",
    "% occupied G&A beds occupied by patients with a length of stay of - 14 or more days",
    "% occupied G&A beds occupied by patients with a length of stay of - 21 or more days",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def extract_period_from_filename(filepath):
    """
    Extract period string from sitrep filename.
    e.g. '202204-April-2022-sitrep...' -> '2022-04-01'
    e.g. '202603-March-2026-beds...' -> '2026-03-01'
    Returns ISO date string or None.
    """
    basename = os.path.basename(filepath)
    match = re.match(r"(\d{4})(\d{2})", basename)
    if match:
        return f"{match.group(1)}-{match.group(2)}-01"
    return None


def find_all_acute_sheet(wb):
    """Find the 'all acute trusts' sheet in a sitrep workbook."""
    for name in wb.sheetnames:
        if "all acute" in name.lower():
            return name
    return None


# ── Sitrep Excel loader (April 2022 to March 2024) ───────────────────────────

import re as _re
_TRUST_CODE_PATTERN = _re.compile(r'^R[A-Z0-9]{2}[A-Z0-9]?$')


def strip_quotes(value):
    """
    Strip surrounding single quotes from a cell value.
    Some NHS Excel files store text cells as 'value' with literal quote characters.
    e.g. "'RC9'" -> "RC9",  "'Org Code'" -> "Org Code"
    Returns the cleaned string, or None if value is None.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s.startswith("'") and s.endswith("'") and len(s) >= 2:
        return s[1:-1].strip()
    return s


def find_trust_data_start(ws):
    """
    Find the first row containing a valid NHS trust org code (starts with R).
    Works across all sitrep file formats:
      - Standard (Apr 2022 to Jul 2023): org code in col 3, data from row 27
      - Revised (Aug 2023 to Mar 2024):  org code in col 2, data from row 70
      - April 2023: no header row, org code in col 3, data from row 27
    NHS trust codes start with R (e.g. RC9, RGT, R0A).
    ICB codes start with Q and are skipped.
    Returns (first_data_row, org_code_col) or (None, None).
    """
    for i, row in enumerate(ws.iter_rows(min_row=20, max_row=200, values_only=True), start=20):
        for col_idx in [2, 3]:
            cell = row[col_idx] if len(row) > col_idx else None
            if cell:
                s = strip_quotes(cell)
                if s and _TRUST_CODE_PATTERN.match(s):
                    return i, col_idx
    return None, None


def load_sitrep_excel(filepath, period_str):
    """
    Load trust-level data from a sitrep Excel file.
    Dynamically finds the trust-level header row (works for both standard
    and revised formats which have different layouts).
    Standard format: header at row 26, org code in column D (index 3)
    Revised format:  header at row 69, org code in column C (index 2)
    Returns a cleaned DataFrame or None.
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = find_all_acute_sheet(wb)
        if not sheet_name:
            print(f"  WARNING: No 'all acute' sheet found in {os.path.basename(filepath)}")
            wb.close()
            return None

        ws = wb[sheet_name]

        # Find first trust data row using NHS org code pattern (starts with R)
        first_data_row, org_code_col = find_trust_data_start(ws)
        if first_data_row is None:
            print(f"  WARNING: Could not find trust data rows in {os.path.basename(filepath)}")
            wb.close()
            return None

        # Read the metric header from row 15 which is consistent across all formats.
        # The row immediately above trust data only works for standard files;
        # revised files (Aug 2023 onwards) have ICB data above trust rows so
        # the metric headers remain at row 15 throughout.
        header_row = list(ws.iter_rows(min_row=15, max_row=15, values_only=True))[0]

        # Read all data rows from first_data_row onwards
        data_rows = []
        for row in ws.iter_rows(min_row=first_data_row, values_only=True):
            if any(v is not None for v in row):
                data_rows.append(row)

        wb.close()

        if not data_rows:
            print(f"  WARNING: No data rows found in {os.path.basename(filepath)}")
            return None

        # Build DataFrame
        df = pd.DataFrame(data_rows, columns=range(len(header_row)))

        # Column mapping depends on format:
        # Standard (Code at col 3): col1=Region, col2=Trust Name, col3=Org Code
        # Revised  (Org Code at col 2): col1=Region, col2=Org Code, col3=Org Name
        if org_code_col == 3:
            # Standard format
            df = df.rename(columns={1: "region", 2: "org_name", 3: "org_code"})
            metric_start = 4
        else:
            # Revised format: org_code_col == 2
            df = df.rename(columns={1: "region", 2: "org_code", 3: "org_name"})
            metric_start = 4

        # Drop the empty first column
        df = df.drop(columns=[0], errors="ignore")

        # Build metric column names from header row
        # Apply strip_quotes to handle files where headers are stored as 'value'
        # Skip duplicate names and '-' placeholder values to avoid DataFrame errors
        metric_col_map = {}
        seen_metric_names = set()
        for i, col_name in enumerate(header_row):
            if i >= metric_start and col_name is not None:
                cleaned = strip_quotes(col_name)
                if cleaned and cleaned != '-' and cleaned not in seen_metric_names:
                    metric_col_map[i] = cleaned
                    seen_metric_names.add(cleaned)

        df = df.rename(columns=metric_col_map)

        # Strip quotes from org_code, org_name, and region values
        for col in ["org_code", "org_name", "region"]:
            if col in df.columns:
                df[col] = df[col].apply(strip_quotes)

        # Drop rows where org_code is missing (these are subtotal/blank rows)
        df = df[df["org_code"].notna()].copy()
        df = df[df["org_code"].astype(str).str.strip() != ""].copy()

        # Add period date
        df["period_date"] = pd.to_datetime(period_str)

        # Convert metric columns to numeric
        # Guard against duplicate column names which return a DataFrame not a Series
        for col in df.columns:
            if col not in ["period_date", "region", "org_name", "org_code"]:
                if isinstance(df[col], pd.Series):
                    df[col] = pd.to_numeric(df[col], errors="coerce")

        # Keep only columns we care about
        keep_fixed = ["period_date", "region", "org_code", "org_name"]
        keep_metrics = [c for c in SITREP_METRICS_TO_KEEP if c in df.columns]
        df = df[keep_fixed + keep_metrics].copy()

        return df

    except Exception as e:
        print(f"  WARNING: Could not process {os.path.basename(filepath)}: {e}")
        return None


# ── Sitrep CSV loader (April 2024 onwards) ───────────────────────────────────

def load_sitrep_csv(filepath, period_str):
    """
    Load trust-level data from a sitrep CSV file.
    Filters to Level = Provider, pivots metrics to wide format.
    Returns a cleaned DataFrame or None.
    """
    try:
        df = pd.read_csv(filepath, dtype=str, low_memory=False)

        # Standardise column names
        df.columns = [c.strip() for c in df.columns]

        # Filter to provider (trust) level only
        if "Level" not in df.columns:
            print(f"  WARNING: No 'Level' column in {os.path.basename(filepath)}")
            return None

        df = df[df["Level"] == "Provider"].copy()

        if len(df) == 0:
            print(f"  WARNING: No Provider-level rows in {os.path.basename(filepath)}")
            return None

        # Keep only metrics we want
        if "Metric" in df.columns:
            df = df[df["Metric"].isin(SITREP_METRICS_TO_KEEP)].copy()

        # Keep Type 1 and All Type rows (drop Other which is non-standard)
        # For TrustPulse we use All Type as the primary metric
        if "Type" in df.columns:
            df = df[df["Type"].isin(["All Type", "Type 1"])].copy()
            # Prefer All Type; use Type 1 only if All Type not available
            # Pivot will handle this by taking the first value per org/metric
            df = df.sort_values("Type", ascending=False)  # All Type sorts after Type 1

        # Convert Value to numeric
        df["Value"] = pd.to_numeric(df["Value"], errors="coerce")

        # Pivot to wide format: one row per trust, one column per metric
        id_cols = ["Org Code", "Org Name", "Region"]
        id_cols = [c for c in id_cols if c in df.columns]

        if "Metric" not in df.columns or "Value" not in df.columns:
            print(f"  WARNING: Missing Metric or Value columns in {os.path.basename(filepath)}")
            return None

        # If Type column present, pivot on Metric for All Type only first
        if "Type" in df.columns:
            df_alltype = df[df["Type"] == "All Type"].copy()
            if len(df_alltype) > 0:
                df = df_alltype

        pivoted = df.pivot_table(
            index=id_cols,
            columns="Metric",
            values="Value",
            aggfunc="first"
        ).reset_index()

        # Flatten column names
        pivoted.columns.name = None

        # Rename to match Excel format column names
        rename_map = {
            "Org Code": "org_code",
            "Org Name": "org_name",
            "Region": "region",
        }
        pivoted = pivoted.rename(columns=rename_map)

        # Add period date
        pivoted["period_date"] = pd.to_datetime(period_str)

        return pivoted

    except Exception as e:
        print(f"  WARNING: Could not process {os.path.basename(filepath)}: {e}")
        return None


# ── KH03 loader ───────────────────────────────────────────────────────────────

def load_kh03_file(filepath):
    """
    Load trust-level data from a KH03 quarterly Excel file.
    Sheet: 'NHS Trust by Sector'
    Headers span rows 14-15. Data starts at row 16.
    Metric groups: Available (cols 6-10), Occupied (cols 12-16), % Occupied (cols 18-22)
    Sectors: Total, General & Acute, Learning Disability/Disabilities, Maternity, Mental Illness
    Returns a cleaned DataFrame or None.
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)

        if "NHS Trust by Sector" not in wb.sheetnames:
            print(f"  WARNING: No 'NHS Trust by Sector' sheet in {os.path.basename(filepath)}")
            wb.close()
            return None

        ws = wb["NHS Trust by Sector"]

        # Extract period from row 5 (Period: October to December 2025)
        period_cell = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
        period_str_raw = None
        for cell in period_cell:
            if cell and isinstance(cell, str) and len(cell) > 5:
                period_str_raw = cell.strip()
                break

        # Extract year from row 5 context and build period_date
        # Period string like "October to December 2025" or "April to June 2023"
        period_date = None
        if period_str_raw:
            # Extract the last month and year mentioned
            months = {
                "January": "01", "February": "02", "March": "03", "April": "04",
                "May": "05", "June": "06", "July": "07", "August": "08",
                "September": "09", "October": "10", "November": "11", "December": "12"
            }
            year_match = re.search(r"(\d{4})", period_str_raw)
            # Get the last month mentioned in the string (end of quarter)
            last_month = None
            for month_name in months:
                if month_name in period_str_raw:
                    last_month = months[month_name]
            if year_match and last_month:
                period_date = pd.to_datetime(f"{year_match.group(1)}-{last_month}-01")

        if period_date is None:
            print(f"  WARNING: Could not parse period from {os.path.basename(filepath)}")
            # Try to extract from filename
            basename = os.path.basename(filepath)
            year_match = re.search(r"(\d{4})", basename)
            if year_match:
                period_date = pd.to_datetime(f"{year_match.group(1)}-01-01")
            else:
                wb.close()
                return None

        # Row 14: group headers (Available, Occupied, % Occupied)
        # Row 15: column names (Year, Period End, Region Code, Org Code, Org Name,
        #                       Total, G&A, Learning Disability, Maternity, Mental Illness x3)
        row14 = list(ws.iter_rows(min_row=14, max_row=14, values_only=True))[0]
        row15 = list(ws.iter_rows(min_row=15, max_row=15, values_only=True))[0]

        # Build column names by combining group header + sector name
        # Positions confirmed from data exploration:
        # Col 1(B): Year, Col 2(C): Period End, Col 3(D): Region Code
        # Col 4(E): Org Code, Col 5(F): Org Name
        # Cols 6-10: Available (Total, G&A, Learning Disability, Maternity, Mental Illness)
        # Col 11: empty
        # Cols 12-16: Occupied (same sectors)
        # Col 17: empty
        # Cols 18-22: % Occupied (same sectors)

        col_names = []
        current_group = None
        for i, (g, c) in enumerate(zip(row14, row15)):
            if g is not None:
                current_group = str(g).strip()
            if c is not None:
                c_clean = str(c).strip().rstrip()
                if current_group and i >= 6:
                    # Normalise learning disability naming
                    c_clean = c_clean.replace("Learning Disabilities", "Learning Disability")
                    col_names.append(f"{current_group}_{c_clean}")
                else:
                    col_names.append(c_clean)
            else:
                col_names.append(None)

        # Read data rows from row 16
        data_rows = []
        for row in ws.iter_rows(min_row=16, values_only=True):
            if any(v is not None for v in row):
                data_rows.append(row)

        wb.close()

        if not data_rows:
            print(f"  WARNING: No data rows in {os.path.basename(filepath)}")
            return None

        df = pd.DataFrame(data_rows, columns=range(len(col_names)))

        # Rename using col_names
        rename_map = {i: name for i, name in enumerate(col_names) if name is not None}
        df = df.rename(columns=rename_map)

        # Drop unnamed/None columns
        df = df[[c for c in df.columns if c is not None and str(c) != "None"]]

        # Keep only rows with a valid Org Code (trust rows, not England/region totals)
        if "Org Code" in df.columns:
            df = df[df["Org Code"].notna()].copy()
            df = df[df["Org Code"].astype(str).str.strip() != ""].copy()
            # Remove England and region-level rows (they have no Org Code but some
            # files may include them with text in the Org Name column)
            df = df[~df["Org Code"].astype(str).str.contains("England|Region", na=False)].copy()

        # Add period date
        df["period_date"] = period_date

        # Convert numeric columns
        for col in df.columns:
            if col not in ["period_date", "Year", "Period End", "Region Code",
                           "Org Code", "Org Name"]:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # Standardise column names
        df = df.rename(columns={
            "Org Code": "org_code",
            "Org Name": "org_name",
            "Region Code": "region_code",
            "Year": "nhs_year",
            "Period End": "period_end_month",
        })

        return df

    except Exception as e:
        print(f"  WARNING: Could not process {os.path.basename(filepath)}: {e}")
        return None


# ── Main ingestion logic ──────────────────────────────────────────────────────

def ingest_beds():

    # ── Part 1: Sitrep ────────────────────────────────────────────────────────
    print("=" * 60)
    print("TrustPulse | Bed Sitrep Ingestion")
    print("=" * 60)

    excel_files = sorted(glob.glob(os.path.join(SITREP_DIR, "*.xlsx")))
    csv_files = sorted(glob.glob(os.path.join(SITREP_DIR, "*.csv")))
    all_sitrep = excel_files + csv_files

    if not all_sitrep:
        print(f"ERROR: No sitrep files found in {SITREP_DIR}")
    else:
        print(f"Found {len(excel_files)} Excel files and {len(csv_files)} CSV files")

        frames = []
        for i, filepath in enumerate(all_sitrep, 1):
            basename = os.path.basename(filepath)
            period_str = extract_period_from_filename(filepath)

            if not period_str:
                print(f"[{i}/{len(all_sitrep)}] SKIPPED (could not parse date): {basename}")
                continue

            print(f"[{i}/{len(all_sitrep)}] {period_str[:7]} — {basename[:55]}...")

            if filepath.endswith(".xlsx"):
                df = load_sitrep_excel(filepath, period_str)
            else:
                df = load_sitrep_csv(filepath, period_str)

            if df is not None and len(df) > 0:
                print(f"  Rows: {len(df):,} | Orgs: {df['org_code'].nunique()}")
                frames.append(df)

        if frames:
            sitrep_combined = pd.concat(frames, ignore_index=True)
            before = len(sitrep_combined)
            sitrep_combined = sitrep_combined.drop_duplicates()
            after = len(sitrep_combined)
            if before != after:
                print(f"Removed {before - after:,} duplicate rows")

            sort_cols = ["period_date", "org_code"]
            sitrep_combined = sitrep_combined.sort_values(sort_cols).reset_index(drop=True)

            os.makedirs(PROCESSED_DIR, exist_ok=True)
            sitrep_combined.to_csv(SITREP_OUTPUT, index=False)
            print(f"\nSaved: {SITREP_OUTPUT}")
            print("\n── Sitrep Summary ───────────────────────────────────────")
            print(f"  Total rows:    {len(sitrep_combined):,}")
            print(f"  Columns:       {sitrep_combined.shape[1]}")
            print(f"  Unique trusts: {sitrep_combined['org_code'].nunique():,}")
            date_min = sitrep_combined["period_date"].min().strftime("%B %Y")
            date_max = sitrep_combined["period_date"].max().strftime("%B %Y")
            print(f"  Date range:    {date_min} to {date_max}")
            print("─────────────────────────────────────────────────────────")

    # ── Part 2: KH03 ─────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("TrustPulse | KH03 Bed Occupancy Ingestion")
    print("=" * 60)

    kh03_files = sorted(glob.glob(os.path.join(KH03_DIR, "*.xlsx")))

    if not kh03_files:
        print(f"ERROR: No KH03 Excel files found in {KH03_DIR}")
    else:
        print(f"Found {len(kh03_files)} KH03 files")

        kh03_frames = []
        for i, filepath in enumerate(kh03_files, 1):
            basename = os.path.basename(filepath)
            print(f"[{i}/{len(kh03_files)}] {basename[:60]}...")
            df = load_kh03_file(filepath)
            if df is not None and len(df) > 0:
                print(f"  Rows: {len(df):,} | Orgs: {df['org_code'].nunique()}")
                kh03_frames.append(df)

        if kh03_frames:
            kh03_combined = pd.concat(kh03_frames, ignore_index=True)
            before = len(kh03_combined)
            kh03_combined = kh03_combined.drop_duplicates()
            after = len(kh03_combined)
            if before != after:
                print(f"Removed {before - after:,} duplicate rows")

            kh03_combined = kh03_combined.sort_values(
                ["period_date", "org_code"]
            ).reset_index(drop=True)

            kh03_combined.to_csv(KH03_OUTPUT, index=False)
            print(f"\nSaved: {KH03_OUTPUT}")
            print("\n── KH03 Summary ─────────────────────────────────────────")
            print(f"  Total rows:    {len(kh03_combined):,}")
            print(f"  Columns:       {kh03_combined.shape[1]}")
            print(f"  Unique trusts: {kh03_combined['org_code'].nunique():,}")
            date_min = kh03_combined["period_date"].min().strftime("%B %Y")
            date_max = kh03_combined["period_date"].max().strftime("%B %Y")
            print(f"  Date range:    {date_min} to {date_max}")
            print("─────────────────────────────────────────────────────────")

    print("\nBed ingestion complete.")


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ingest_beds()
