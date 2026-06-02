"""
TrustPulse — pipeline/ingest/cancelled_ops.py
Ingests NHS Cancelled Elective Operations data from two sources:

1. Quarterly Excel/xlsm files (data/raw/cancelled_ops/Q*.xlsx, Q*.xlsm)
   Sheet: Provider
   Header: row 16
   Columns: Region Code, Region Name, Organisation Code, Organisation Name,
            Number of last minute elective operations cancelled for non clinical reasons,
            Number of patients not treated within 28 days of last minute cancellation
   Trust rows start at row 20 (rows 17-18 are England totals, row 19 blank)
   Period extracted from row 5 of Provider sheet

2. QMCO Annual files (QMCO-Annual-CSV-*.csv and QMCO-Annual-CSV-*.xlsx)
   Columns: Year, Period Name, Parent Org Code, Parent Name,
            Org Code, Org Name, Cancelled Operations, Breaches Of Standard
   Monthly granularity within each NHS year (June, September, December, March)

Outputs:
    data/processed/cancelled_ops_quarterly_clean.csv  — quarterly trust-level data
    data/processed/cancelled_ops_monthly_clean.csv    — monthly trust-level data from QMCO

Usage:
    python pipeline\ingest\cancelled_ops.py
"""

import os, re, glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# ── Paths ─────────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RAW_DIR = os.path.join(BASE_DIR, "data", "raw", "cancelled_ops")
PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
QUARTERLY_OUTPUT = os.path.join(PROCESSED_DIR, "cancelled_ops_quarterly_clean.csv")
MONTHLY_OUTPUT = os.path.join(PROCESSED_DIR, "cancelled_ops_monthly_clean.csv")

_TRUST_CODE = re.compile(r'^R[A-Z0-9]{2}[A-Z0-9]?$')

# Month name to number mapping for QMCO period parsing
MONTH_MAP = {
    'JANUARY': 1, 'FEBRUARY': 2, 'MARCH': 3, 'APRIL': 4,
    'MAY': 5, 'JUNE': 6, 'JULY': 7, 'AUGUST': 8,
    'SEPTEMBER': 9, 'OCTOBER': 10, 'NOVEMBER': 11, 'DECEMBER': 12
}


# ── Quarterly Excel/xlsm loader ───────────────────────────────────────────────

def parse_quarter_period(period_str):
    """
    Parse period string like 'Quarter 1, 2022-23 (April to June 2022)'
    Returns the end date of the quarter as a pandas Timestamp.
    e.g. Q1 -> June, Q2 -> September, Q3 -> December, Q4 -> March
    """
    if not period_str:
        return None
    s = str(period_str).strip()

    # Extract quarter number and year
    q_match = re.search(r'Quarter\s+(\d)', s, re.IGNORECASE)
    # Extract the calendar year from the bracket e.g. "(April to June 2022)"
    year_match = re.search(r'\(.*?(\d{4})\)', s)
    # Also try to get NHS year like "2022-23"
    nhs_year_match = re.search(r'(\d{4})-(\d{2})', s)

    quarter_end_months = {1: 6, 2: 9, 3: 12, 4: 3}

    if q_match and year_match:
        q = int(q_match.group(1))
        year = int(year_match.group(1))
        month = quarter_end_months.get(q, 6)
        # Q4 ends in March of the following year
        if q == 4 and nhs_year_match:
            year = int('20' + nhs_year_match.group(2))
        return pd.Timestamp(year=year, month=month, day=1)

    return None


def load_quarterly_file(filepath):
    """Load one quarterly cancelled ops Excel or xlsm file."""
    try:
        # Try openpyxl first; fall back to pandas for files openpyxl cannot read
        wb = load_workbook(filepath, read_only=True, data_only=True, keep_vba=False)
        if 'Provider' not in wb.sheetnames:
            print(f"  WARNING: No Provider sheet in {os.path.basename(filepath)}")
            wb.close()
            return None
        ws = wb['Provider']
        # Check if openpyxl can actually read the content
        test_rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
        openpyxl_works = any(any(v for v in row) for row in test_rows)
        wb.close()

        if openpyxl_works:
            # Standard openpyxl path
            wb = load_workbook(filepath, read_only=True, data_only=True, keep_vba=False)
            ws = wb['Provider']
            period_date = None
            for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                if len(row) > 2 and row[1] and str(row[1]).strip() == 'Period:':
                    period_date = parse_quarter_period(row[2])
                    break
            if period_date is None:
                print(f"  WARNING: Could not parse period from {os.path.basename(filepath)}")
                wb.close()
                return None
            header = list(ws.iter_rows(min_row=16, max_row=16, values_only=True))[0]
            data_rows = [
                row for row in ws.iter_rows(min_row=17, values_only=True)
                if any(v is not None for v in row)
            ]
            wb.close()
        else:
            # Pandas fallback for files openpyxl cannot read
            import pandas as _pd
            raw = _pd.read_excel(filepath, sheet_name='Provider',
                                 header=None, engine='openpyxl')
            period_date = None
            for _, row in raw.iterrows():
                vals = row.tolist()
                if len(vals) > 2 and str(vals[1]).strip() == 'Period:':
                    period_date = parse_quarter_period(vals[2])
                    break
            if period_date is None:
                print(f"  WARNING: Could not parse period from {os.path.basename(filepath)}")
                return None
            # Header is at row index 15 (0-based), data from row 16
            header = tuple(raw.iloc[15].tolist())
            data_rows = [tuple(r) for _, r in raw.iloc[16:].iterrows()
                         if any(v is not None and str(v) != 'nan' for v in r)]

        if not data_rows:
            return None

        # Build column names
        cols = []
        for v in header:
            if v is None:
                cols.append('__DROP__')
            else:
                c = re.sub(r'[^a-z0-9]+', '_', str(v).strip().lower()).strip('_')
                cols.append(c if c else '__DROP__')

        df = pd.DataFrame(data_rows, columns=cols)
        df = df[[c for c in df.columns if c != '__DROP__']]

        # Rename columns
        rename = {
            'region_code': 'region_code',
            'region_name': 'region_name',
            'organisation_code': 'org_code',
            'organisation_name': 'org_name',
        }
        # Handle long column names for the two metrics
        for col in df.columns:
            if 'last_minute' in col and 'cancelled' in col:
                rename[col] = 'num_cancelled'
            elif 'not_treated' in col or '28_days' in col or 'patients_not' in col:
                rename[col] = 'num_not_treated_28_days'
        df = df.rename(columns=rename)

        # Filter to trust rows only
        if 'org_code' not in df.columns:
            return None
        df = df[df['org_code'].apply(
            lambda x: bool(_TRUST_CODE.match(str(x).strip())) if x else False
        )].copy()

        if len(df) == 0:
            return None

        # Convert numeric
        for col in ['num_cancelled', 'num_not_treated_28_days']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        df['period_date'] = period_date
        df['data_source'] = 'quarterly_excel'

        keep = ['period_date', 'region_code', 'region_name', 'org_code', 'org_name',
                'num_cancelled', 'num_not_treated_28_days', 'data_source']
        return df[[c for c in keep if c in df.columns]].copy()

    except Exception as e:
        import traceback
        print(f"  WARNING: {os.path.basename(filepath)}: {e}")
        traceback.print_exc()
        return None


# ── QMCO Annual loader ────────────────────────────────────────────────────────

def parse_qmco_period(year_str, period_name):
    """
    Parse QMCO period into a Timestamp.
    year_str: '2022-23' or '2025 - 26'
    period_name: 'JUNE', 'SEPTEMBER', 'DECEMBER', 'MARCH'
    Returns Timestamp of the first day of the period month.
    """
    try:
        # Normalise year string
        year_clean = re.sub(r'\s', '', str(year_str))  # '2022-23' or '2025-26'
        start_year = int(year_clean[:4])
        end_year = int('20' + year_clean[-2:])
        month = MONTH_MAP.get(str(period_name).strip().upper())
        if not month:
            return None
        # March belongs to the end year (e.g. March 2023 for 2022-23)
        year = end_year if month <= 6 else start_year
        return pd.Timestamp(year=year, month=month, day=1)
    except Exception:
        return None


def load_qmco_csv(filepath):
    """Load one QMCO Annual CSV file."""
    try:
        df = pd.read_csv(filepath, dtype=str)
        return df
    except Exception as e:
        print(f"  WARNING: {os.path.basename(filepath)}: {e}")
        return None


def load_qmco_xlsx(filepath):
    """Load one QMCO Annual XLSX file."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return None
        header = [str(v).strip() if v else f'col_{i}' for i, v in enumerate(rows[0])]
        data = rows[1:]
        return pd.DataFrame(data, columns=header)
    except Exception as e:
        print(f"  WARNING: {os.path.basename(filepath)}: {e}")
        return None


def process_qmco_df(df, source_file):
    """Clean and standardise a QMCO dataframe."""
    # Standardise column names
    df.columns = [re.sub(r'[^a-z0-9]+', '_', str(c).strip().lower()).strip('_')
                  for c in df.columns]

    rename = {
        'year': 'nhs_year',
        'period_name': 'period_name',
        'parent_org_code': 'parent_org_code',
        'parent_name': 'parent_name',
        'org_code': 'org_code',
        'org_name': 'org_name',
        'cancelled_operations': 'num_cancelled',
        'breaches_of_standard': 'num_not_treated_28_days',
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # Filter to trust rows only
    if 'org_code' not in df.columns:
        return None
    df = df[df['org_code'].apply(
        lambda x: bool(_TRUST_CODE.match(str(x).strip())) if x else False
    )].copy()

    if len(df) == 0:
        return None

    # Parse period date
    if 'nhs_year' in df.columns and 'period_name' in df.columns:
        df['period_date'] = df.apply(
            lambda r: parse_qmco_period(r['nhs_year'], r['period_name']), axis=1
        )
    else:
        return None

    df = df[df['period_date'].notna()].copy()

    # Convert numeric
    for col in ['num_cancelled', 'num_not_treated_28_days']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    df['data_source'] = 'qmco_annual'

    keep = ['period_date', 'nhs_year', 'period_name', 'parent_org_code', 'parent_name',
            'org_code', 'org_name', 'num_cancelled', 'num_not_treated_28_days', 'data_source']
    return df[[c for c in keep if c in df.columns]].copy()


# ── Main ingestion logic ──────────────────────────────────────────────────────

def ingest_cancelled_ops():
    print("=" * 60)
    print("TrustPulse | Cancelled Operations Ingestion")
    print("=" * 60)

    os.makedirs(PROCESSED_DIR, exist_ok=True)

    # ── Part 1: Quarterly Excel/xlsm files ───────────────────────────────────
    print("\n── Part 1: Quarterly Excel files ────────────────────────")
    q_files = sorted(
        glob.glob(os.path.join(RAW_DIR, "Q*.xlsx")) +
        glob.glob(os.path.join(RAW_DIR, "Q*.xlsm"))
    )
    print(f"Found {len(q_files)} quarterly files")

    q_frames = []
    for i, fp in enumerate(q_files, 1):
        print(f"[{i}/{len(q_files)}] {os.path.basename(fp)[:55]}...")
        df = load_quarterly_file(fp)
        if df is not None and len(df) > 0:
            print(f"  Rows: {len(df):,} | Orgs: {df['org_code'].nunique()} | Period: {df['period_date'].iloc[0].strftime('%b %Y')}")
            q_frames.append(df)
        else:
            print(f"  SKIPPED")

    if q_frames:
        q_combined = pd.concat(q_frames, ignore_index=True)
        q_combined = q_combined.drop_duplicates()
        q_combined = q_combined.sort_values(['period_date', 'org_code']).reset_index(drop=True)
        q_combined.to_csv(QUARTERLY_OUTPUT, index=False)
        print(f"\nSaved: {QUARTERLY_OUTPUT}")
        print(f"  Rows: {len(q_combined):,} | Trusts: {q_combined['org_code'].nunique()} | "
              f"Range: {q_combined['period_date'].min().strftime('%b %Y')} to "
              f"{q_combined['period_date'].max().strftime('%b %Y')}")

    # ── Part 2: QMCO Annual files ─────────────────────────────────────────────
    print("\n── Part 2: QMCO Annual files ────────────────────────────")
    qmco_csv = sorted(glob.glob(os.path.join(RAW_DIR, "QMCO*.csv")))
    # Only pick QMCO xlsx files, not quarterly Q* files
    qmco_xlsx = sorted([f for f in glob.glob(os.path.join(RAW_DIR, "QMCO*.xlsx"))])
    qmco_files = qmco_csv + qmco_xlsx
    print(f"Found {len(qmco_files)} QMCO files ({len(qmco_csv)} CSV, {len(qmco_xlsx)} XLSX)")

    m_frames = []
    for i, fp in enumerate(qmco_files, 1):
        print(f"[{i}/{len(qmco_files)}] {os.path.basename(fp)[:55]}...")
        raw = load_qmco_csv(fp) if fp.endswith('.csv') else load_qmco_xlsx(fp)
        if raw is None:
            print(f"  SKIPPED")
            continue
        df = process_qmco_df(raw, fp)
        if df is not None and len(df) > 0:
            print(f"  Rows: {len(df):,} | Orgs: {df['org_code'].nunique()} | "
                  f"Periods: {sorted(df['period_date'].dt.strftime('%b %Y').unique())}")
            m_frames.append(df)
        else:
            print(f"  SKIPPED")

    if m_frames:
        m_combined = pd.concat(m_frames, ignore_index=True)
        m_combined = m_combined.drop_duplicates()
        m_combined = m_combined.sort_values(['period_date', 'org_code']).reset_index(drop=True)
        m_combined.to_csv(MONTHLY_OUTPUT, index=False)
        print(f"\nSaved: {MONTHLY_OUTPUT}")
        print(f"  Rows: {len(m_combined):,} | Trusts: {m_combined['org_code'].nunique()} | "
              f"Range: {m_combined['period_date'].min().strftime('%b %Y')} to "
              f"{m_combined['period_date'].max().strftime('%b %Y')}")

    print("\n── Summary ──────────────────────────────────────────────")
    if q_frames:
        print(f"  Quarterly output: {len(q_combined):,} rows, {q_combined.shape[1]} cols")
    if m_frames:
        print(f"  QMCO output:      {len(m_combined):,} rows, {m_combined.shape[1]} cols")
    print("─────────────────────────────────────────────────────────")
    print("Cancelled ops ingestion complete.")


if __name__ == "__main__":
    ingest_cancelled_ops()
