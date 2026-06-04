"""
pipeline/ingest/ambulance.py
TrustPulse -- Ambulance Handover Delays Ingest
Source: NHS England UEC Sitrep, Web File Timeseries Ambulance Collection
URL: https://www.england.nhs.uk/statistics/statistical-work-areas/uec-sitrep/
     urgent-and-emergency-care-daily-situation-reports-2025-26/
Coverage: 24 November 2025 to 29 March 2026 (daily, 126 days, 149 trusts)

File structure:
  Sheet: Handovers
  Row 3 (0-indexed row 2): dates, one per 14-column block, starting at col index 3
  Row 5 (0-indexed row 4): column sub-headers per date block (14 cols each)
  Row 8 (0-indexed row 7): England aggregate -- skipped
  Row 9 (0-indexed row 8): blank -- skipped
  Rows 10+ (0-indexed row 9+): trust-level data

Column layout within each 14-column date block (0-based offset):
  0  Handover time known (count)
  1  Over 15 minutes (count)
  2  Over 30 minutes (count)
  3  Over 45 minutes (count)   -- captured but not used downstream
  4  Over 60 minutes (count)
  5  Handover time unknown (count)
  6  All handovers (count)     -- = known + unknown
  7  Total hours
  8  Mean (as timedelta)
  9  Proportion over 15 min
  10 Proportion over 30 min
  11 Proportion over 45 min
  12 Proportion over 60 min
  13 Proportion unknown

Output: data/processed/ambulance_clean.csv
Columns: org_code, trust_name, region, date (YYYY-MM-01), and monthly aggregates:
  amb_handovers_total       -- sum of all_handovers across days in month
  amb_handovers_known       -- sum of handover_time_known
  amb_over15_count          -- sum of over_15_min
  amb_over30_count          -- sum of over_30_min
  amb_over60_count          -- sum of over_60_min
  amb_over15_pct            -- amb_over15_count / amb_handovers_known (where known > 0)
  amb_over30_pct            -- amb_over30_count / amb_handovers_known
  amb_over60_pct            -- amb_over60_count / amb_handovers_known

Notes:
  - Proportions are recalculated from counts rather than averaging the daily
    proportion columns, to avoid distortion from days with zero handovers.
  - Daily data is aggregated to calendar month (first of month as date key).
  - Trusts with zero handovers for a full month retain a row with NaN proportions.
  - Date column uses period_start = first day of calendar month, consistent with
    other TrustPulse ingest scripts.
"""

import os
import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.join(os.path.dirname(__file__), "..", "..")
RAW_DIR  = os.path.join(BASE_DIR, "data", "raw", "ambulance")
OUT_DIR  = os.path.join(BASE_DIR, "data", "processed")
OUT_FILE = os.path.join(OUT_DIR, "ambulance_clean.csv")

# ---------------------------------------------------------------------------
# File discovery -- find the first .xlsx in the raw/ambulance folder
# ---------------------------------------------------------------------------
def find_source_file(raw_dir: str) -> str:
    xlsx_files = [
        f for f in os.listdir(raw_dir)
        if f.lower().endswith(".xlsx") or f.lower().endswith(".xls")
    ]
    if not xlsx_files:
        raise FileNotFoundError(
            f"No Excel file found in {raw_dir}. "
            "Download the 'Web File Timeseries -- Ambulance Collection' from "
            "https://www.england.nhs.uk/statistics/statistical-work-areas/"
            "uec-sitrep/urgent-and-emergency-care-daily-situation-reports-2025-26/ "
            "and save it to data/raw/ambulance/"
        )
    if len(xlsx_files) > 1:
        print(f"  [WARNING] Multiple files found in {raw_dir}; using: {xlsx_files[0]}")
    return os.path.join(raw_dir, xlsx_files[0])


# ---------------------------------------------------------------------------
# Column offsets within each 14-column date block (0-based)
# ---------------------------------------------------------------------------
OFF_KNOWN  = 0   # handover time known
OFF_15     = 1   # over 15 min
OFF_30     = 2   # over 30 min
OFF_60     = 4   # over 60 min (index 3 = over 45 min, skipped)
OFF_ALL    = 6   # all handovers (known + unknown)

DATE_STRIDE  = 14   # columns per date block
DATE_START_COL = 3  # first date block starts at this 0-based column index

SHEET_NAME   = "Handovers"
DATE_ROW     = 3    # 1-based row number containing dates
HEADER_ROW   = 5    # 1-based row number containing column sub-headers
DATA_ROW_START = 10 # 1-based first trust data row (row 8 = England, row 9 = blank)


def run():
    print("[ambulance] Starting ambulance handover delay ingest...")

    src = find_source_file(RAW_DIR)
    print(f"  Source file : {os.path.basename(src)}")

    # ------------------------------------------------------------------
    # Load workbook in read_only mode
    # ------------------------------------------------------------------
    wb = load_workbook(src, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    # ------------------------------------------------------------------
    # Extract dates from row 3
    # Each date appears at column index DATE_START_COL + n * DATE_STRIDE
    # ------------------------------------------------------------------
    all_rows = list(ws.iter_rows(values_only=True))

    date_row   = all_rows[DATE_ROW - 1]        # 0-indexed
    dates = []
    col = DATE_START_COL
    while col < len(date_row):
        val = date_row[col]
        if val is not None:
            # openpyxl returns datetime for date cells
            if isinstance(val, pd.Timestamp) or hasattr(val, "date"):
                dates.append(pd.Timestamp(val).normalize())
            else:
                dates.append(pd.Timestamp(str(val)).normalize())
        col += DATE_STRIDE

    n_dates = len(dates)
    print(f"  Date range  : {dates[0].date()} to {dates[-1].date()} ({n_dates} days)")

    # ------------------------------------------------------------------
    # Parse trust rows
    # Each trust row: col[0]=region, col[1]=org_code, col[2]=trust_name
    # Then for each date block d (0-based): base_col = DATE_START_COL + d * DATE_STRIDE
    # ------------------------------------------------------------------
    records = []
    skipped = 0

    for row in all_rows[DATA_ROW_START - 1:]:
        org_code   = row[1]
        trust_name = row[2]
        region     = row[0]

        # Skip aggregate rows (England), blank rows, and any header remnants
        if org_code is None or trust_name is None:
            skipped += 1
            continue
        if not isinstance(org_code, str) or len(org_code.strip()) == 0:
            skipped += 1
            continue

        org_code   = org_code.strip()
        trust_name = str(trust_name).strip()
        region     = str(region).strip() if region else ""

        for d_idx, date in enumerate(dates):
            base = DATE_START_COL + d_idx * DATE_STRIDE

            def _val(offset):
                """Safely extract numeric value; return None if missing or non-numeric."""
                try:
                    v = row[base + offset]
                    if v is None:
                        return None
                    return float(v)
                except (TypeError, ValueError, IndexError):
                    return None

            records.append({
                "org_code"   : org_code,
                "trust_name" : trust_name,
                "region"     : region,
                "_date"      : date,
                "known"      : _val(OFF_KNOWN),
                "over15"     : _val(OFF_15),
                "over30"     : _val(OFF_30),
                "over60"     : _val(OFF_60),
                "all_hov"    : _val(OFF_ALL),
            })

    print(f"  Trust rows  : {len(records) // n_dates} trusts x {n_dates} days "
          f"= {len(records)} daily records")
    if skipped:
        print(f"  [INFO] Skipped {skipped} blank/aggregate rows")

    # ------------------------------------------------------------------
    # Build daily DataFrame
    # ------------------------------------------------------------------
    df = pd.DataFrame(records)
    df["_date"] = pd.to_datetime(df["_date"])

    # ------------------------------------------------------------------
    # Aggregate to calendar month
    # period_start = first day of the calendar month (consistent with
    # other TrustPulse ingest scripts which use YYYY-MM-01)
    # ------------------------------------------------------------------
    df["date"] = df["_date"].dt.to_period("M").dt.to_timestamp()

    monthly = (
        df.groupby(["org_code", "trust_name", "region", "date"])
        .agg(
            amb_handovers_total = ("all_hov",  "sum"),
            amb_handovers_known = ("known",    "sum"),
            amb_over15_count    = ("over15",   "sum"),
            amb_over30_count    = ("over30",   "sum"),
            amb_over60_count    = ("over60",   "sum"),
        )
        .reset_index()
    )

    # ------------------------------------------------------------------
    # Recalculate proportions from monthly counts
    # Use amb_handovers_known as denominator (matches NHS published method)
    # ------------------------------------------------------------------
    def safe_pct(numerator, denominator):
        return numerator.where(denominator > 0).div(
            denominator.where(denominator > 0)
        )

    monthly["amb_over15_pct"] = safe_pct(
        monthly["amb_over15_count"], monthly["amb_handovers_known"]
    )
    monthly["amb_over30_pct"] = safe_pct(
        monthly["amb_over30_count"], monthly["amb_handovers_known"]
    )
    monthly["amb_over60_pct"] = safe_pct(
        monthly["amb_over60_count"], monthly["amb_handovers_known"]
    )

    # ------------------------------------------------------------------
    # Sort and output
    # ------------------------------------------------------------------
    monthly.sort_values(["org_code", "date"], inplace=True)
    monthly.reset_index(drop=True, inplace=True)

    os.makedirs(OUT_DIR, exist_ok=True)
    monthly.to_csv(OUT_FILE, index=False)

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    n_trusts = monthly["org_code"].nunique()
    n_months = monthly["date"].nunique()
    null_pct  = monthly["amb_over15_pct"].isna().mean() * 100

    print(f"  Output      : {OUT_FILE}")
    print(f"  Rows        : {len(monthly):,}  ({n_trusts} trusts x {n_months} months)")
    print(f"  Date range  : {monthly['date'].min().date()} to {monthly['date'].max().date()}")
    print(f"  amb_over15_pct null : {null_pct:.1f}%")
    print("[ambulance] Done.")
    return monthly


if __name__ == "__main__":
    run()
