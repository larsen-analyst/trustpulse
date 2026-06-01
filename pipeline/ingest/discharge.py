"""
TrustPulse — pipeline/ingest/discharge.py
Ingests NHS Discharge Ready Date monthly Excel files.
"""

import os, re, glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RAW_DIR = os.path.join(BASE_DIR, "data", "raw", "discharge_delays")
PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
OUTPUT_FILE = os.path.join(PROCESSED_DIR, "discharge_clean.csv")

_TRUST_CODE = re.compile(r'^R[A-Z0-9]{2}[A-Z0-9]?$')

DELAY_BANDS = {
    '1 day', '2-3 days', '4-6 days', '7-13 days', '14-20 days', '21 days or more',
    'no delay', '1 day delay', '2-3 day delay', '4-6 day delay',
    '7-13 day delay', '14-20 day delay',
}

def extract_period(ws):
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        if row[0] and str(row[0]).strip() == 'Period:':
            try:
                return pd.to_datetime(row[1])
            except Exception:
                pass
    return None

def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        vals = [str(v).strip() if v else '' for v in row]
        if 'Org Code' in vals or ('Code' in vals and 'Organisation Name' in vals):
            return i
    return None

def build_cols(header):
    occurrence = {}
    col_names = []
    for v in header:
        if not v or str(v).strip() in ('', '-'):
            col_names.append('__DROP__')
            continue
        raw = str(v).strip().lower()
        c = re.sub(r'[^a-z0-9]+', '_', raw).strip('_')
        is_band = any(d in raw for d in DELAY_BANDS)
        if is_band:
            count = occurrence.get(c, 0)
            occurrence[c] = count + 1
            prefixes = ['acc', 'unacc', 'acc_count', 'unacc_count']
            col_names.append(f"{prefixes[min(count, 3)]}_{c}")
        else:
            if c not in occurrence:
                occurrence[c] = 1
                col_names.append(c)
            else:
                occurrence[c] += 1
                col_names.append(f'{c}_{occurrence[c]}')
    return col_names

def standardise(df):
    rename = {
        'code': 'org_code',
        'organisation_name': 'org_name',
        'number_of_providers_submitting_acceptable_data': 'num_providers_submitting',
        'of_providers_submitting_acceptable_data': 'pct_providers_submitting',
        'number_of_patients_discharged_in_total': 'num_patients_discharged',
        'total_bed_days_lost_due_to_delayed_discharge': 'total_bed_days_lost',
        'date_of_discharge_is_same_as_discharge_ready_date': 'pct_same_day',
        'date_of_discharge_is_1_days_after_discharge_ready_date': 'pct_1plus_days',
        'average_days_from_discharge_ready_date_to_date_of_discharge_exc_0_day_delays': 'avg_days_exc_0',
        'average_days_from_discharge_ready_date_to_date_of_discharge_inc_0_day_delays': 'avg_days_inc_0',
        'average_days_from_discharge_ready_date_to_date_of_discharge': 'avg_days_exc_0',
    }
    return df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

def load_file(filepath):
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        if 'Provider' not in wb.sheetnames:
            wb.close()
            return None
        ws = wb['Provider']
        period = extract_period(ws)
        if not period:
            wb.close()
            return None
        hr = find_header_row(ws)
        if not hr:
            wb.close()
            return None
        header = list(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))[0]
        cols = build_cols(header)
        data_rows = [r for r in ws.iter_rows(min_row=hr+1, values_only=True)
                     if any(v is not None for v in r)]
        wb.close()
        if not data_rows:
            return None

        df = pd.DataFrame(data_rows, columns=cols)

        # Drop blank/separator columns — use string name __DROP__ to avoid NaN issues
        df = df[[c for c in df.columns if c != '__DROP__']]

        # Filter to trust rows
        org_col = 'org_code' if 'org_code' in df.columns else 'code'
        if org_col not in df.columns:
            return None
        df = df[df[org_col].apply(
            lambda x: bool(_TRUST_CODE.match(str(x).strip())) if x else False
        )].copy()
        if len(df) == 0:
            return None

        if org_col == 'code':
            df = df.rename(columns={'code': 'org_code'})

        df = standardise(df)
        df['period_date'] = period

        skip = {'period_date', 'org_code', 'org_name', 'region', 'icb', 'data_source'}
        for col in df.columns:
            if col not in skip and isinstance(df[col], pd.Series):
                df[col] = pd.to_numeric(df[col], errors='coerce')

        return df

    except Exception as e:
        import traceback
        print(f"  WARNING: {os.path.basename(filepath)}: {e}")
        traceback.print_exc()
        return None

def ingest_discharge():
    print("=" * 60)
    print("TrustPulse | Discharge Delays Ingestion")
    print("=" * 60)

    files = sorted(glob.glob(os.path.join(RAW_DIR, "*.xlsx")))
    if not files:
        print(f"ERROR: No files in {RAW_DIR}")
        return

    print(f"Found {len(files)} files")
    frames = []
    skipped = 0

    for i, fp in enumerate(files, 1):
        print(f"[{i}/{len(files)}] {os.path.basename(fp)[:60]}...")
        df = load_file(fp)
        if df is not None and len(df) > 0:
            print(f"  Rows: {len(df):,} | Orgs: {df['org_code'].nunique()}")
            frames.append(df)
        else:
            skipped += 1

    if not frames:
        print("ERROR: No files processed.")
        return

    combined = pd.concat(frames, ignore_index=True)
    print(f"\nCombined: {combined.shape[0]:,} rows x {combined.shape[1]} columns")

    before = len(combined)
    combined = combined.drop_duplicates()
    if before != len(combined):
        print(f"Removed {before - len(combined):,} duplicates")

    combined = combined.sort_values(['period_date', 'org_code']).reset_index(drop=True)
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    combined.to_csv(OUTPUT_FILE, index=False)
    print(f"\nSaved: {OUTPUT_FILE}")

    print("\n── Summary ──────────────────────────────────────────────")
    print(f"  Total rows:    {len(combined):,}")
    print(f"  Columns:       {combined.shape[1]}")
    print(f"  Unique trusts: {combined['org_code'].nunique():,}")
    print(f"  Date range:    {combined['period_date'].min().strftime('%B %Y')} to {combined['period_date'].max().strftime('%B %Y')}")
    if 'total_bed_days_lost' in combined.columns:
        latest = combined['period_date'].max()
        lost = combined[combined['period_date'] == latest]['total_bed_days_lost'].sum()
        if lost > 0:
            print(f"  Bed days lost ({latest.strftime('%B %Y')}): {lost:,.0f}")
            print(f"  Cost at £345/day: £{lost * 345:,.0f}")
            print(f"  NOTE: £345/day is NHS England published rate.")
    if skipped:
        print(f"  Skipped: {skipped} files")
    print("─────────────────────────────────────────────────────────")
    print("Discharge ingestion complete.")

if __name__ == "__main__":
    ingest_discharge()
